import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

# 定义请求头，模拟浏览器行为
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
    'Referer': 'https://www.example.com/'
}

# 目标URL
url = 'https://tianqi.2345.com/Pc/GetHistory?areaInfo%5BareaId%5D=59488&areaInfo%5BareaType%5D=2&date%5Byear%5D=2024&date%5Bmonth%5D=2'

try:
    # 发送 GET 请求
    response = requests.get(url, headers=headers)
    response.raise_for_status()  # 检查请求是否成功

    # 解析 JSON 数据
    weather_data = response.json()

    # 检查 JSON 数据中是否包含嵌入的HTML
    if 'data' in weather_data:
        # 提取嵌入的HTML字符串
        embedded_html = weather_data['data']

        # 解析HTML内容
        soup = BeautifulSoup(embedded_html, 'html.parser')

        # 找到历史天气数据的表格
        weather_table = soup.find('table', {'class': 'history-table'})

        if weather_table:
            # 使用 pandas 读取表格数据
            df = pd.read_html(str(weather_table))[0]

            # 创建一个新的 Excel 工作簿
            wb = Workbook()
            ws = wb.active

            # 写入表头
            for col_num, column_title in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_num, value=column_title)

            # 写入数据并设置样式
            for row_num, row_data in enumerate(df.itertuples(index=False), 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                    # 设置样式
                    if isinstance(cell_value, str):
                        if '°' in cell_value:
                            if 'color:#ff5040;' in weather_table.find_all('td')[col_num - 1].get('style', ''):
                                cell.font = Font(color="FF0000")
                            elif 'color:#3097fd;' in weather_table.find_all('td')[col_num - 1].get('style', ''):
                                cell.font = Font(color="0000FF")
                        elif '优' in cell_value:
                            cell.font = Font(color="00FF00")
                        elif '良' in cell_value:
                            cell.font = Font(color="FFFF00")

            # 保存到 Excel 文件
            wb.save('weather_data_styled.xlsx')
            print("已将数据写入到 weather_data_styled.xlsx 文件中")
        else:
            print("未找到历史天气数据表格")

    else:
        print("未找到历史天气数据")

except requests.exceptions.RequestException as e:
    print(f"请求失败: {e}")

except json.JSONDecodeError as e:
    print(f"JSON解析失败: {e}")
